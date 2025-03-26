import feedparser
import re
import openpyxl
import os.path
from datetime import datetime


acronym_name = {
    # "NĐ"   : "Nghị định",
    # "TT"   : "Thông tư",
    # "CĐ"   : "Công điện",
    # "QĐ"   : "Quyết định",
    "BTTTT": "Bộ Thông tin và Truyền thông",
    "BKHCN": "Bộ Khoa học và Công nghệ",
    "TTg"  : "Thủ tướng Chính phủ",
    "CP"   : "Chính phủ",
    "BCT"  : "Bộ Công Thương",
    "VPQH" : "Văn phòng Quốc hội",
    "VPCTN" : "Văn phòng Chủ tịch nước",
    "VPCP" : "Văn phòng Chính phủ",
    "TANDTC" : "Tòa án nhân dân tối cao",
    "VKSNDTC" : "Viện Kiểm sát ND tối cao",
    "TTCP" : "Thanh tra Chính phủ"
}

ministry_shortname = {
    "BCA" : "Bộ Công an",
    "BNG" : "Bộ Ngoại giao",
    "BTP" : "Bộ Tư pháp",
    "KTNN" : "Kiểm toán Nhà nước",
    "BKHĐT" : "Bộ Kế hoạch và Đầu tư",
    "HLHPNVN" : "Hội LH Phụ nữ Việt Nam",
    "ĐTNCSHCM" : "TW Đoàn TN CS HCM",
    "MTTQ" : "UB TW MTTQ Việt Nam",
    "LMHTX" : "Liên minh HTX Việt Nam",
    "HND" : "Hội Nông dân Việt Nam",
    "HCCB" : "Hội Cựu chiến binh Việt Nam",
    "BNV" : "Bộ Nội vụ",
    "BTC" : "Bộ Tài chính",
    "BVHTTDL" : "Bộ Văn hóa- Thể thao- Du lịch",
    "BGDĐT" : "Bộ Giáo dục và Đào tạo",
    "QGHN" : "Đại học Quốc gia Hà nội",
    "QGHCM" : "Đại học Quốc gia TP. HCM",
    "BKHCN" : "Bộ Khoa học và Công nghệ",
    "TTXVN" : "Thông tấn xã Việt nam",
    "THVN" : "Đài Truyền hình Việt nam",
    "TNVN" : "Đài Tiếng nói Việt nam",
    "KHCNVN" : "Viện Khoa học và công nghệ VN",
    "KHXHVN" : "Viện Khoa học xã hội ViệtNam",
    "BQLHL" : "BQL khu c.nghệ cao Hoà Lạc",
    "BQLVHDL" : "Ban QL làng văn hoá Du lịch",
    "BYT" : "Bộ Y tế",
    "LĐLĐVN" : "Tổng Liên đoàn LĐ Việt Nam",
    "BLĐTBXH" : "Bộ Lao động - TB&XH",
    "BHXHVN" : "Bảo hiểm xã hội Việt nam",
    "UBDT" : "Uỷ ban Dân tộc",
    "BNN" : "Bộ Nông nghiệp và PTNT",
    "BCT" : "Bộ Công thương",
    "BTNMT" : "Bộ Tài nguyên - Môi trường",
    "BXD" : "Bộ Xây dựng",
    "BGTVT" : "Bộ Giao thông vận tải",
    "UBSMC" : "Uỷ ban Sông Mê Kông",
    "TCT" : "Tổng cục Thuế",
    "CT" : "Cục thuế",
    "CHQ" : "Cục Hải quan"
}

include_word = [
    "Công nghệ thông tin",
    "Truyền thông",
    "Cổng thông tin điện tử",
    "Chuyển đổi số",
    "An toàn thông tin",
    "Mạng",
    "Ứng dụng công nghệ thông tin",
    "Hệ thống thông tin",
    "Cơ sở dữ liệu",
    "Dữ liệu",
    "Điện tử",
    "Bưu chính",
    "Kết nối và chia sẻ dữ liệu",
    "Dịch vụ Internet",
    "Viễn thông",
    "Mạng truyền số liệu chuyên dùng",
    "Báo",
    "Báo chí",
    "Xuất bản",
    "Thông tin và Truyền thông",
    "Di động",
    "Công nghệ",
    "Internet",
    "Phần mềm",
    "Thông tin",
    "Chữ ký điện tử",
    "Sở hữu trí tuệ",
    "Nhiệm vụ khoa học và công nghệ",
    "Sáng kiến",
    "Đổi mới sáng tạo",
    "Chuyển giao công nghệ",
    "An toàn bức xạ"
]


exclude_word = [
    "/QĐ-UBND",
    "NQ-HĐND",
    "KH-UBND",
    "UBND-KSTTHC",
    "Dự thảo"
]

exclude_type = [
    "Quy chuẩn"
]

excel = ""
url = "https://thuvienphapluat.vn/rss.xml"
docs = feedparser.parse(url)

# Step 1: Filter data
def filter_docs(title):
    key_word = []

    if "BTTTT" in title or acronym_name["BTTTT"] in title:
        return True
    if any(word in title for word in exclude_word):
        return False
    if any(word in title for word in ministry_shortname):
        return False
    for word in include_word:
        if word in title:
            key_word.append(word)
    if key_word:
        return True

    return False

def filter_type(type_name):
    if any(word in type_name for word in exclude_type):
        return False
    
    return True

def date_format(date):
    datetime_object = datetime.strptime(date, '%a, %d %b %Y %H:%M:%S %Z')
    formatted_date = datetime_object.strftime('%d/%m/%Y')

    return formatted_date


def extract_numbers(title):
    pattern = r'\b\d{1,7}/[\w-]+(?:/[\w-]+)?'
    matches = re.findall(pattern, title)
    return matches[0] if matches else ""


# def ministry_name(doc_number):
#     extracted_string = extract_numbers(doc_number)
#     if '-' in extracted_string:
#         return extracted_string.split('-')[-1]
#     else:
#         return ""


def ministry_name(doc_number):
    return next((element for element in acronym_name if element in doc_number), None)


# Step 2: Add excel
def remove_pattern_from_string(line):
    pattern = r'\b\d{1,4}/[\w-]+(?:/[\w-]+)?'
    match = re.search(pattern, line)
    
    if match:
        return line.replace(match.group(), "", 1).replace("  ", " ")
    return line

def get_month():
    month = datetime.now().month
    day = datetime.now().day
    if day > 20:
        return month + 1
    else:
        return month

def main():

    if not os.path.isfile("./BC Tháng " + str(get_month()) + ".xlsx"):
        excel = "./Template.xlsx"
    else:
        excel = "./BC Tháng " + str(get_month()) + ".xlsx"


    # 1. Filter data
    docs_list = []
    for doc in docs.entries:
        doc_filtered = []
        if filter_docs(doc.title) and filter_type(doc.tags[0].term[9:-3]):
            type_name = doc.tags[0].term[9:-3]
            doc_filtered.append(type_name)
            doc_filtered.append(extract_numbers(doc.title))
            doc_filtered.append(date_format(doc.published))
            doc_filtered.append(ministry_name(extract_numbers(doc.title)))
            doc_filtered.append(remove_pattern_from_string(doc.title))
            doc_filtered.append(doc.links[0].href)
        else:
            continue
        docs_list.append(doc_filtered)

    docs_list = sorted(docs_list, key=lambda x: datetime.strptime(x[2], "%d/%m/%Y"))
    # Reverse order sort
    # docs_list = docs_list[::-1]
    # print(docs_list)
    # val = input("Stop!!!")

    # 2. Remove duplicate data
    file_path = 'BC Tháng ' + str(get_month()) + '.xlsx'
    docs_filtered = docs_list
    if os.path.isfile(file_path):

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        data = []

        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row[0] is not None:  # Check if the cell is not empty
                data.append(extract_numbers(row[0]))

        docs_filtered = [sublist for sublist in docs_list if sublist[1] not in data]

    if docs_filtered:
        print("Có " + str(len(docs_filtered)) + " văn bản mới:")
        for doc in docs_filtered:
            print(doc)
            print("Số: " + doc[1] + " ngày " + doc[2] + " của " + acronym_name[doc[3]])
    else:
        print("Không có văn bản mới!")


    # 3. Add to excel
    wb = openpyxl.load_workbook(excel)
    ws = wb.active

    thin_border = openpyxl.styles.Border(
        left = openpyxl.styles.Side(style="thin"),
        right = openpyxl.styles.Side(style="thin"),
        top = openpyxl.styles.Side(style="thin"),
        bottom = openpyxl.styles.Side(style="thin")
    )
    
    for row_index, row_value in enumerate(docs_filtered, start = ws.max_row + 1):
        ws.cell(row=row_index, column=1, value=row_index - 1)
        ws.cell(row=row_index, column=2, value=row_value[0])
        ws.cell(row=row_index, column=3, value="Số: " + row_value[1] + " ngày " + row_value[2] + " của " + acronym_name[row_value[3]])
        ws.cell(row=row_index, column=4, value=remove_pattern_from_string(row_value[4]))
        ws.cell(row=row_index, column=6, value=remove_pattern_from_string(row_value[4]))
        ws.cell(row=row_index, column=6).value='=HYPERLINK("{}", "{}")'.format(row_value[5], "Link Name")

        # Add border for a row
        for col in range(1,7):
            cell = ws.cell(row=row_index, column=col)
            cell.border = thin_border

    wrap_text_alignment = openpyxl.styles.Alignment(vertical='center', wrapText=True)
    center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_text_alignment

    for cell in ws['A']:
        cell.alignment = center_alignment

    for cell in ws[1]:
        cell.alignment = center_alignment

    wb.save("BC Tháng " + str(get_month()) + ".xlsx")

    


if __name__ == "__main__":
    main()


# print(docs.entries[0].title)
# print(docs.entries[0].link)
# print(date_format(docs.entries[0].published))
# print(docs.entries[0].tags[0].term[9:-3])